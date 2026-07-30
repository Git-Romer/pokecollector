"""Review and import John John's PC collection workbooks.

The review pass is deliberately read-only. The commit pass reruns the same
validation and applies the resulting operations in one transaction.
"""

from __future__ import annotations

import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models import Card, CollectionItem, ProductPurchase, StorageLocation
from services.inventory import (
    ACQUISITION_SOURCES,
    CARD_VARIANTS,
    PROTECTION_TYPES,
    RAW_CONDITIONS,
    REMOVAL_REASONS,
    SEALED_CONDITIONS,
    changed_values,
    get_or_create_default_storage_location,
    normalize_card_variant,
    record_inventory_event,
)
from services.tcgdex_languages import is_supported_tcgdex_language, normalize_tcgdex_language


REQUIRED_SHEETS = (
    "Owned Cards",
    "Bulk",
    "Sealed Products",
    "Storage Locations",
    "Import Errors",
)
MAX_WORKBOOK_BYTES = 10 * 1024 * 1024

CARD_FIELDS = (
    "card_id",
    "quantity",
    "condition",
    "variant",
    "lang",
    "purchase_price",
    "acquisition_source",
    "collection_intent",
    "inventory_kind",
    "protection_type",
    "storage_location_id",
    "grader",
    "grade",
    "certification_number",
    "is_grail",
    "card_history",
    "notes",
    "status",
    "removal_reason",
)
PRODUCT_FIELDS = (
    "product_name",
    "product_type",
    "acquisition_source",
    "collection_intent",
    "quantity",
    "sealed_condition",
    "purchase_price",
    "purchase_date",
    "storage_location_id",
    "notes",
    "status",
    "removal_reason",
)
LOCATION_FIELDS = ("name", "description", "is_default", "is_active")


class WorkbookValueError(ValueError):
    pass


def _text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    return str(value).strip()


def _optional_text(value: Any) -> str | None:
    normalized = _text(value)
    return normalized or None


def _positive_int(value: Any, field: str) -> int:
    try:
        normalized = int(value)
    except (TypeError, ValueError):
        raise WorkbookValueError(f"{field} must be a whole number")
    if normalized < 1:
        raise WorkbookValueError(f"{field} must be at least 1")
    return normalized


def _optional_float(value: Any, field: str) -> float | None:
    if value in (None, ""):
        return None
    try:
        normalized = float(value)
    except (TypeError, ValueError):
        raise WorkbookValueError(f"{field} must be a number")
    if normalized < 0:
        raise WorkbookValueError(f"{field} cannot be negative")
    return normalized


def _bool(value: Any, default: bool) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    normalized = _text(value).lower()
    if normalized in {"true", "yes", "1", "active"}:
        return True
    if normalized in {"false", "no", "0", "inactive"}:
        return False
    raise WorkbookValueError("Value must be True or False")


def _date(value: Any, field: str) -> datetime.date:
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    try:
        return datetime.date.fromisoformat(_text(value))
    except ValueError:
        raise WorkbookValueError(f"{field} must use YYYY-MM-DD")


def _rows(sheet):
    headers = [_text(cell.value) for cell in sheet[1]]
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in cells]
        if not any(value not in (None, "") for value in values):
            continue
        yield row_number, dict(zip(headers, values))


def _state(record, fields):
    return {field: getattr(record, field) for field in fields}


def _find_owned_record(db, model, user_id: int, record_uid: str | None):
    if not record_uid:
        return None
    record = db.query(model).filter(model.record_uid == record_uid).first()
    if record is not None and record.user_id != user_id:
        raise WorkbookValueError("Record UID belongs to another user")
    return record


def _add_error(result, sheet: str, row: int, record_uid: str | None, message: str):
    result["errors"].append(
        {
            "sheet": sheet,
            "row": row,
            "record_uid": record_uid or "",
            "error": message,
        }
    )


def _location_key(uid: str | None, name: str | None) -> str:
    return f"uid:{uid}" if uid else f"name:{(name or '').casefold()}"


def _resolve_location_operation(location_ops, uid: str | None, name: str | None):
    if uid:
        match = next((op for op in location_ops if op["record_uid"] == uid), None)
        if match:
            return match
    if name:
        folded = name.casefold()
        return next((op for op in location_ops if op["values"]["name"].casefold() == folded), None)
    return None


def _classify(result, operation, before: dict | None, after: dict):
    if before is None:
        action = "create"
        result["summary"]["new_records"] += 1
    elif changed_values(before, after):
        action = "update"
        result["summary"]["updated_records"] += 1
    else:
        action = "unchanged"
        result["summary"]["unchanged_records"] += 1
    operation["action"] = action
    result["actions"].append(
        {
            "sheet": operation["sheet"],
            "row": operation["row"],
            "record_uid": operation["record_uid"] or "",
            "action": action,
            "name": operation["name"],
        }
    )


def _plan_locations(db: Session, user_id: int, workbook, result):
    operations = []
    seen = set()
    for row_number, row in _rows(workbook["Storage Locations"]):
        uid = _optional_text(row.get("Record UID"))
        name = _text(row.get("Name"))
        try:
            if not name:
                raise WorkbookValueError("Name is required")
            key = _location_key(uid, name)
            if key in seen:
                result["summary"]["duplicates"] += 1
                raise WorkbookValueError("Duplicate storage location in workbook")
            seen.add(key)
            record = _find_owned_record(db, StorageLocation, user_id, uid)
            if record is None:
                record = db.query(StorageLocation).filter(
                    StorageLocation.user_id == user_id,
                    StorageLocation.name == name,
                ).first()
            values = {
                "name": name,
                "description": _optional_text(row.get("Description")),
                "is_default": _bool(row.get("Default"), False),
                "is_active": _bool(row.get("Active"), True),
            }
            operation = {
                "sheet": "Storage Locations",
                "row": row_number,
                "record_uid": uid,
                "name": name,
                "record": record,
                "values": values,
            }
            _classify(result, operation, _state(record, LOCATION_FIELDS) if record else None, values)
            operations.append(operation)
        except WorkbookValueError as exc:
            _add_error(result, "Storage Locations", row_number, uid, str(exc))
    return operations


def _plan_cards(db: Session, user_id: int, workbook, result, location_ops):
    operations = []
    seen = set()
    for sheet_name, inventory_kind in (("Owned Cards", "owned"), ("Bulk", "bulk")):
        for row_number, row in _rows(workbook[sheet_name]):
            uid = _optional_text(row.get("Record UID"))
            card_id = _text(row.get("Card ID"))
            try:
                if uid and uid in seen:
                    result["summary"]["duplicates"] += 1
                    raise WorkbookValueError("Duplicate card Record UID in workbook")
                if uid:
                    seen.add(uid)
                if not card_id:
                    raise WorkbookValueError("Card ID is required")
                card = db.query(Card).filter(Card.id == card_id).first()
                if card is None:
                    raise WorkbookValueError(
                        "Card ID is not in the local catalog; sync or add the exact printing first"
                    )
                result["summary"]["matched_cards"] += 1

                condition = _text(row.get("Condition"), "NM")
                if condition not in RAW_CONDITIONS:
                    raise WorkbookValueError(f"Condition must be one of: {', '.join(RAW_CONDITIONS)}")
                protection = _text(row.get("Protection"), "raw")
                if protection not in PROTECTION_TYPES:
                    raise WorkbookValueError(
                        f"Protection must be one of: {', '.join(PROTECTION_TYPES)}"
                    )
                source = _optional_text(row.get("Acquisition Source"))
                if source and source not in ACQUISITION_SOURCES:
                    raise WorkbookValueError(
                        f"Acquisition Source must be one of: {', '.join(ACQUISITION_SOURCES)}"
                    )
                collection_intent = _text(row.get("Collection Intent"), "main_collection")
                if collection_intent not in {"main_collection", "vault", "pc"}:
                    raise WorkbookValueError("Collection Intent must be main_collection, vault, or pc")
                status = _text(row.get("Status"), "owned")
                if status not in {"owned", "removed"}:
                    raise WorkbookValueError("Status must be owned or removed")
                removal_reason = _optional_text(row.get("Removal Reason"))
                if removal_reason and removal_reason not in REMOVAL_REASONS:
                    raise WorkbookValueError(
                        f"Removal Reason must be one of: {', '.join(REMOVAL_REASONS)}"
                    )

                location_uid = _optional_text(row.get("Storage Location UID"))
                location_name = _optional_text(row.get("Storage Location"))
                location_op = _resolve_location_operation(location_ops, location_uid, location_name)
                if location_op is None and (location_uid or location_name):
                    raise WorkbookValueError("Storage location is not listed in Storage Locations")

                cost = _optional_float(row.get("Cost Basis"), "Cost Basis")
                if inventory_kind == "bulk":
                    cost = None
                elif cost is None and source == "pulled":
                    cost = 4.49
                grader = _optional_text(row.get("Grading Company"))
                if protection == "psa_slab" and not grader:
                    grader = "PSA"

                record = _find_owned_record(db, CollectionItem, user_id, uid)
                variant = normalize_card_variant(_text(row.get("Variant"), "Normal"))
                if variant not in CARD_VARIANTS:
                    raise WorkbookValueError(f"Variant must be one of: {', '.join(CARD_VARIANTS)}")
                language = normalize_tcgdex_language(_text(row.get("Language"), card.lang or "en"))
                if not is_supported_tcgdex_language(language):
                    raise WorkbookValueError("Language is not a supported TCGdex language code")
                if language != (card.lang or "en"):
                    raise WorkbookValueError("Language must match the exact Card ID printing")

                values = {
                    "card_id": card_id,
                    "quantity": _positive_int(row.get("Quantity"), "Quantity"),
                    "condition": condition,
                    "variant": variant,
                    "lang": language,
                    "purchase_price": cost,
                    "acquisition_source": source,
                    "collection_intent": collection_intent,
                    "inventory_kind": inventory_kind,
                    "protection_type": protection,
                    "storage_location_id": location_op["record"].id if location_op and location_op["record"] else None,
                    "grader": grader,
                    "grade": _optional_text(row.get("Grade")),
                    "certification_number": _optional_text(row.get("Certification Number")),
                    "is_grail": _bool(row.get("Grail"), False),
                    "card_history": _optional_text(row.get("Card History")),
                    "notes": _optional_text(row.get("Notes")),
                    "status": status,
                    "removal_reason": removal_reason,
                }
                before = _state(record, CARD_FIELDS) if record else None
                comparable = dict(values)
                if location_op and location_op["record"] is None:
                    comparable["storage_location_id"] = f"pending:{_location_key(location_uid, location_name)}"
                    if before:
                        before["storage_location_id"] = record.storage_location_id
                operation = {
                    "sheet": sheet_name,
                    "row": row_number,
                    "record_uid": uid,
                    "name": card.name,
                    "record": record,
                    "values": values,
                    "location_op": location_op,
                }
                _classify(result, operation, before, comparable)
                operations.append(operation)
            except WorkbookValueError as exc:
                _add_error(result, sheet_name, row_number, uid, str(exc))
    return operations


def _plan_products(db: Session, user_id: int, workbook, result, location_ops):
    operations = []
    seen = set()
    for row_number, row in _rows(workbook["Sealed Products"]):
        uid = _optional_text(row.get("Record UID"))
        name = _text(row.get("Product Name"))
        try:
            if uid and uid in seen:
                result["summary"]["duplicates"] += 1
                raise WorkbookValueError("Duplicate sealed-product Record UID in workbook")
            if uid:
                seen.add(uid)
            if not name:
                raise WorkbookValueError("Product Name is required")
            condition = _text(row.get("Condition"), "factory_sealed")
            if condition not in SEALED_CONDITIONS:
                raise WorkbookValueError(
                    f"Condition must be one of: {', '.join(SEALED_CONDITIONS)}"
                )
            collection_intent = _text(row.get("Collection Intent"), "main_collection")
            if collection_intent not in {"main_collection", "vault", "pc"}:
                raise WorkbookValueError("Collection Intent must be main_collection, vault, or pc")
            status = _text(row.get("Status"), "active")
            if status not in {"active", "removed"}:
                raise WorkbookValueError("Status must be active or removed")
            removal_reason = _optional_text(row.get("Removal Reason"))
            if removal_reason and removal_reason not in REMOVAL_REASONS:
                raise WorkbookValueError(
                    f"Removal Reason must be one of: {', '.join(REMOVAL_REASONS)}"
                )
            acquisition_source = _optional_text(row.get("Acquisition Source"))
            if acquisition_source and acquisition_source not in ACQUISITION_SOURCES:
                raise WorkbookValueError(
                    f"Acquisition Source must be one of: {', '.join(ACQUISITION_SOURCES)}"
                )
            location_uid = _optional_text(row.get("Storage Location UID"))
            location_name = _optional_text(row.get("Storage Location"))
            location_op = _resolve_location_operation(location_ops, location_uid, location_name)
            if location_op is None and (location_uid or location_name):
                raise WorkbookValueError("Storage location is not listed in Storage Locations")

            record = _find_owned_record(db, ProductPurchase, user_id, uid)
            values = {
                "product_name": name,
                "product_type": _optional_text(row.get("Product Type")),
                "acquisition_source": acquisition_source,
                "collection_intent": collection_intent,
                "quantity": _positive_int(row.get("Quantity"), "Quantity"),
                "sealed_condition": condition,
                "purchase_price": _optional_float(row.get("Cost Basis"), "Cost Basis"),
                "purchase_date": _date(row.get("Acquisition Date"), "Acquisition Date"),
                "storage_location_id": location_op["record"].id if location_op and location_op["record"] else None,
                "notes": _optional_text(row.get("Notes")),
                "status": status,
                "removal_reason": removal_reason,
            }
            if values["purchase_price"] is None:
                raise WorkbookValueError("Cost Basis is required for sealed products")
            before = _state(record, PRODUCT_FIELDS) if record else None
            comparable = dict(values)
            if location_op and location_op["record"] is None:
                comparable["storage_location_id"] = f"pending:{_location_key(location_uid, location_name)}"
            operation = {
                "sheet": "Sealed Products",
                "row": row_number,
                "record_uid": uid,
                "name": name,
                "record": record,
                "values": values,
                "location_op": location_op,
            }
            _classify(result, operation, before, comparable)
            operations.append(operation)
        except WorkbookValueError as exc:
            _add_error(result, "Sealed Products", row_number, uid, str(exc))
    return operations


def _apply_locations(db, user_id, operations):
    for operation in operations:
        record = operation["record"]
        if operation["action"] == "create":
            record = StorageLocation(
                user_id=user_id,
                record_uid=operation["record_uid"] or None,
                **operation["values"],
            )
            db.add(record)
            db.flush()
            operation["record"] = record
            record_inventory_event(
                db,
                user_id=user_id,
                entity_type="storage_location",
                entity_id=record.id,
                entity_uid=record.record_uid,
                action="import_added",
                changes={key: {"before": None, "after": value} for key, value in operation["values"].items()},
            )
        elif operation["action"] == "update":
            before = _state(record, LOCATION_FIELDS)
            for field, value in operation["values"].items():
                setattr(record, field, value)
            record.updated_at = datetime.datetime.utcnow()
            record_inventory_event(
                db,
                user_id=user_id,
                entity_type="storage_location",
                entity_id=record.id,
                entity_uid=record.record_uid,
                action="import_updated",
                changes=changed_values(before, _state(record, LOCATION_FIELDS)),
            )
    defaults = [op["record"] for op in operations if op["values"]["is_default"]]
    if defaults:
        selected = defaults[-1]
        db.query(StorageLocation).filter(
            StorageLocation.user_id == user_id,
            StorageLocation.id != selected.id,
        ).update({"is_default": False}, synchronize_session=False)
    else:
        get_or_create_default_storage_location(db, user_id)
    db.flush()


def _apply_entities(db, user_id, operations, model, entity_type, fields):
    for operation in operations:
        if operation["action"] == "unchanged":
            continue
        values = dict(operation["values"])
        location_op = operation.get("location_op")
        if location_op is not None:
            values["storage_location_id"] = location_op["record"].id
        elif not values.get("storage_location_id"):
            values["storage_location_id"] = get_or_create_default_storage_location(db, user_id).id

        record = operation["record"]
        if operation["action"] == "create":
            record = model(
                user_id=user_id,
                record_uid=operation["record_uid"] or None,
                **values,
            )
            db.add(record)
            db.flush()
            record_inventory_event(
                db,
                user_id=user_id,
                entity_type=entity_type,
                entity_id=record.id,
                entity_uid=record.record_uid,
                action="import_added",
                changes={key: {"before": None, "after": value} for key, value in values.items()},
            )
        else:
            before = _state(record, fields)
            for field, value in values.items():
                setattr(record, field, value)
            record.updated_at = datetime.datetime.utcnow()
            record_inventory_event(
                db,
                user_id=user_id,
                entity_type=entity_type,
                entity_id=record.id,
                entity_uid=record.record_uid,
                action="import_updated",
                changes=changed_values(before, _state(record, fields)),
            )


def review_inventory_workbook(
    db: Session,
    user_id: int,
    payload: bytes,
    *,
    commit: bool = False,
) -> dict:
    """Return an import plan, optionally committing it after complete validation."""

    result = {
        "valid": False,
        "committed": False,
        "summary": {
            "matched_cards": 0,
            "new_records": 0,
            "updated_records": 0,
            "unchanged_records": 0,
            "duplicates": 0,
            "errors": 0,
        },
        "actions": [],
        "errors": [],
    }
    if not payload:
        result["errors"].append({"sheet": "", "row": 0, "record_uid": "", "error": "Workbook is empty"})
        result["summary"]["errors"] = 1
        return result
    if len(payload) > MAX_WORKBOOK_BYTES:
        result["errors"].append(
            {"sheet": "", "row": 0, "record_uid": "", "error": "Workbook exceeds the 10 MB limit"}
        )
        result["summary"]["errors"] = 1
        return result

    try:
        workbook = load_workbook(BytesIO(payload), data_only=True)
    except Exception:
        result["errors"].append(
            {"sheet": "", "row": 0, "record_uid": "", "error": "File is not a readable Excel workbook"}
        )
        result["summary"]["errors"] = 1
        return result

    missing = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
    if missing:
        result["errors"].append(
            {
                "sheet": "",
                "row": 0,
                "record_uid": "",
                "error": f"Missing required sheets: {', '.join(missing)}",
            }
        )
        result["summary"]["errors"] = 1
        return result

    location_ops = _plan_locations(db, user_id, workbook, result)
    card_ops = _plan_cards(db, user_id, workbook, result, location_ops)
    product_ops = _plan_products(db, user_id, workbook, result, location_ops)
    result["summary"]["errors"] = len(result["errors"])
    result["valid"] = not result["errors"]

    if commit and result["valid"]:
        try:
            _apply_locations(db, user_id, location_ops)
            _apply_entities(
                db,
                user_id,
                card_ops,
                CollectionItem,
                "collection_item",
                CARD_FIELDS,
            )
            _apply_entities(
                db,
                user_id,
                product_ops,
                ProductPurchase,
                "sealed_product",
                PRODUCT_FIELDS,
            )
            db.commit()
            result["committed"] = True
        except Exception:
            db.rollback()
            raise

    return result
