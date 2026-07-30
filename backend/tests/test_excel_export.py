import unittest
import datetime
from io import BytesIO
from types import SimpleNamespace

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from api.auth import get_current_user
from api.export import CARD_HEADERS, SEALED_HEADERS, build_collection_workbook, router as export_router
from database import Base, get_db
from models import Card, CollectionItem, ProductPurchase, User


class ExcelExportEndpointTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.db = sessionmaker(bind=self.engine)()
        self.user = User(username="john", hashed_password="x", role="admin", is_active=True)
        other_user = User(username="other", hashed_password="x", role="trainer", is_active=True)
        self.db.add_all([self.user, other_user])
        self.db.flush()

        john_card = Card(
            id="sv1-1_en",
            tcg_card_id="sv1-1",
            name="John's Card",
            number="1",
            lang="en",
        )
        other_card = Card(
            id="sv1-2_en",
            tcg_card_id="sv1-2",
            name="Other User's Card",
            number="2",
            lang="en",
        )
        self.db.add_all([john_card, other_card])
        self.db.flush()
        self.db.add_all([
            CollectionItem(
                user_id=self.user.id,
                card_id=john_card.id,
                quantity=1,
                condition="NM",
                variant="Normal",
                lang="en",
            ),
            CollectionItem(
                user_id=other_user.id,
                card_id=other_card.id,
                quantity=1,
                condition="NM",
                variant="Normal",
                lang="en",
            ),
            ProductPurchase(
                user_id=self.user.id,
                product_name="John's ETB",
                quantity=1,
                purchase_date=datetime.date(2026, 7, 1),
            ),
            ProductPurchase(
                user_id=other_user.id,
                product_name="Other User's ETB",
                quantity=1,
                purchase_date=datetime.date(2026, 7, 1),
            ),
        ])
        self.db.commit()
        self.db.refresh(self.user)

        app = FastAPI()
        app.include_router(export_router, prefix="/api/export")
        app.dependency_overrides[get_db] = lambda: self.db
        app.dependency_overrides[get_current_user] = lambda: self.user
        self.client = TestClient(app)

    def tearDown(self):
        self.client.close()
        self.db.close()
        self.engine.dispose()

    def test_excel_export_has_required_sheets_and_current_user_data(self):
        response = self.client.get("/api/export/xlsx")

        self.assertEqual(response.status_code, 200, response.text)
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames,
            ["Cards", "Sealed Product", "Acquisition & Storage"],
        )
        self.assertEqual(workbook["Cards"]["A1"].value, "Card ID")
        card_names = [cell.value for cell in workbook["Cards"]["B"][1:]]
        product_names = [cell.value for cell in workbook["Sealed Product"]["A"][1:]]
        self.assertIn("John's Card", card_names)
        self.assertNotIn("Other User's Card", card_names)
        self.assertIn("John's ETB", product_names)
        self.assertNotIn("Other User's ETB", product_names)


class ExcelExportTests(unittest.TestCase):
    def test_export_headers_use_cost_basis_language(self):
        self.assertIn("Cost Basis", CARD_HEADERS)
        self.assertIn("Cost Basis", SEALED_HEADERS)
        self.assertNotIn("Purchase Price", CARD_HEADERS)
        self.assertNotIn("Purchase Price", SEALED_HEADERS)

    def test_workbook_contains_required_collection_sheets(self):
        card = SimpleNamespace(
            id="sv1-1_en", name="Test Card", number="1", rarity="Rare",
            set_ref=SimpleNamespace(name="Test Set"),
        )
        item = SimpleNamespace(
            id=7, record_uid="card-record-7", inventory_kind="owned", status="owned",
            card=card, quantity=1, condition="NM", variant="Normal", lang="en",
            purchase_price=4.49, acquisition_source="pulled", storage_type="Penny Sleeve",
            storage_detail="Binder 1", grader=None, grade=None, certification_number=None,
            protection_type="penny_sleeve", collection_intent="pc", is_grail=True,
            card_history="Pulled from a booster pack",
            storage_location=SimpleNamespace(record_uid="location-1", name="Binder 1"),
            notes="First pull", added_at=None, updated_at=None, removal_reason=None,
        )
        product = SimpleNamespace(
            id=3, record_uid="product-record-3", quantity=1, sealed_condition="factory_sealed",
            status="active", removal_reason=None,
            product_name="Elite Trainer Box", product_type="Elite Trainer Box",
            acquisition_source="purchased", collection_intent="vault", purchase_price=None,
            purchase_date=None, storage_type="Sealed", storage_detail="Shelf A",
            storage_location=SimpleNamespace(record_uid="location-2", name="Shelf A"),
            notes=None, updated_at=None,
        )
        locations = [
            SimpleNamespace(
                record_uid="location-1",
                name="Binder 1",
                description="Main binder",
                is_default=False,
                is_active=True,
                created_at=None,
                updated_at=None,
            ),
            SimpleNamespace(
                record_uid="location-2",
                name="Shelf A",
                description="Sealed shelf",
                is_default=False,
                is_active=True,
                created_at=None,
                updated_at=None,
            ),
        ]

        workbook = load_workbook(BytesIO(build_collection_workbook([item], [product], locations)))

        self.assertEqual(
            workbook.sheetnames,
            ["Owned Cards", "Bulk", "Sealed Products", "Storage Locations", "Import Errors"],
        )
        self.assertEqual(workbook["Owned Cards"]["A1"].value, "Record UID")
        self.assertEqual(workbook["Owned Cards"]["A2"].value, "card-record-7")
        self.assertEqual(workbook["Owned Cards"]["M2"].value, "pc")
        self.assertEqual(workbook["Owned Cards"]["T2"].value, True)
        self.assertEqual(workbook["Owned Cards"]["U2"].value, "Pulled from a booster pack")
        self.assertEqual(workbook["Sealed Products"]["A2"].value, "product-record-3")
        self.assertEqual(workbook["Sealed Products"]["E2"].value, "vault")
        self.assertIsNone(workbook["Sealed Products"]["H2"].value)
        self.assertEqual(workbook["Storage Locations"]["A2"].value, "location-1")
        self.assertEqual(workbook["Import Errors"]["A1"].value, "Sheet")
