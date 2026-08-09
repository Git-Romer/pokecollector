import datetime
import unittest
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from api.export import build_collection_workbook
from database import Base
from models import Card, CollectionItem, InventoryEvent, ProductPurchase, StorageLocation, User
from services.inventory_workbook import review_inventory_workbook


class InventoryWorkbookImportTests(unittest.TestCase):
    def setUp(self):
        engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(engine)
        self.db = sessionmaker(bind=engine)()
        self.user = User(username="john", hashed_password="x", role="admin", is_active=True)
        self.card = Card(
            id="sv1-1_en",
            tcg_card_id="sv1-1",
            name="Test Card",
            number="1",
            lang="en",
        )
        self.db.add_all([self.user, self.card])
        self.db.flush()
        self.location = StorageLocation(
            user_id=self.user.id,
            name="Binder 1",
            is_default=True,
        )
        self.db.add(self.location)
        self.db.flush()
        self.item = CollectionItem(
            user_id=self.user.id,
            card_id=self.card.id,
            quantity=1,
            condition="NM",
            variant="Normal",
            lang="en",
            acquisition_source="pulled",
            purchase_price=4.49,
            inventory_kind="owned",
            protection_type="penny_sleeve",
            storage_location_id=self.location.id,
            collection_intent="main_collection",
            is_grail=False,
            card_history=None,
            status="owned",
        )
        self.product = ProductPurchase(
            user_id=self.user.id,
            product_name="Elite Trainer Box",
            product_type="ETB",
            acquisition_source="purchased",
            collection_intent="main_collection",
            quantity=1,
            sealed_condition="factory_sealed",
            purchase_price=49.99,
            purchase_date=datetime.date(2026, 7, 1),
            storage_location_id=self.location.id,
            status="active",
        )
        self.db.add_all([self.item, self.product])
        self.db.commit()
        self.db.refresh(self.item)
        self.db.refresh(self.product)

    def tearDown(self):
        self.db.close()

    def _edited_workbook(self):
        data = build_collection_workbook(
            [self.item],
            [self.product],
            [self.location],
        )
        workbook = load_workbook(BytesIO(data))
        workbook["Owned Cards"]["G2"] = 3
        workbook["Owned Cards"]["M2"] = "pc"
        workbook["Owned Cards"]["T2"] = True
        workbook["Owned Cards"]["U2"] = "Pulled from a booster pack"
        workbook["Owned Cards"]["V2"] = "Moved after review"
        workbook["Sealed Products"]["F2"] = 2
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def test_review_is_read_only_and_commit_updates_by_stable_uid(self):
        data = self._edited_workbook()

        review = review_inventory_workbook(
            self.db,
            self.user.id,
            data,
            commit=False,
        )

        self.assertTrue(review["valid"])
        self.assertEqual(review["summary"]["updated_records"], 2)
        self.assertEqual(self.db.query(CollectionItem).count(), 1)
        self.assertEqual(self.db.get(CollectionItem, self.item.id).quantity, 1)

        committed = review_inventory_workbook(
            self.db,
            self.user.id,
            data,
            commit=True,
        )

        self.assertTrue(committed["committed"])
        self.assertEqual(self.db.query(CollectionItem).count(), 1)
        self.assertEqual(self.db.query(ProductPurchase).count(), 1)
        self.assertEqual(self.db.get(CollectionItem, self.item.id).quantity, 3)
        updated_item = self.db.get(CollectionItem, self.item.id)
        self.assertEqual(updated_item.collection_intent, "pc")
        self.assertTrue(updated_item.is_grail)
        self.assertEqual(updated_item.card_history, "Pulled from a booster pack")
        self.assertEqual(updated_item.notes, "Moved after review")
        self.assertEqual(self.db.get(ProductPurchase, self.product.id).quantity, 2)
        self.assertEqual(
            self.db.query(InventoryEvent).filter(InventoryEvent.action == "import_updated").count(),
            2,
        )

        second = review_inventory_workbook(
            self.db,
            self.user.id,
            data,
            commit=True,
        )
        self.assertTrue(second["committed"])
        self.assertEqual(second["summary"]["unchanged_records"], 3)
        self.assertEqual(self.db.query(CollectionItem).count(), 1)
        self.assertEqual(self.db.query(ProductPurchase).count(), 1)

    def _workbook_with_card_edits(self, **edits):
        data = build_collection_workbook(
            [self.item],
            [self.product],
            [self.location],
        )
        workbook = load_workbook(BytesIO(data))
        for cell, value in edits.items():
            workbook["Owned Cards"][cell] = value
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()


    def test_review_accepts_sealed_product_without_cost_basis(self):
        data = build_collection_workbook(
            [self.item],
            [self.product],
            [self.location],
        )
        workbook = load_workbook(BytesIO(data))
        workbook["Sealed Products"]["H2"] = None
        output = BytesIO()
        workbook.save(output)

        committed = review_inventory_workbook(
            self.db,
            self.user.id,
            output.getvalue(),
            commit=True,
        )

        self.assertTrue(committed["committed"])
        self.assertIsNone(self.db.get(ProductPurchase, self.product.id).purchase_price)

    def test_review_rejects_non_physical_variants(self):
        review = review_inventory_workbook(
            self.db,
            self.user.id,
            self._workbook_with_card_edits(I2="Digital"),
            commit=False,
        )

        self.assertFalse(review["valid"])
        self.assertFalse(review["committed"])
        self.assertTrue(any("Variant must be one of" in error["error"] for error in review["errors"]))
        self.assertEqual(self.db.get(CollectionItem, self.item.id).variant, "Normal")

    def test_review_rejects_language_that_does_not_match_card_printing(self):
        review = review_inventory_workbook(
            self.db,
            self.user.id,
            self._workbook_with_card_edits(J2="de"),
            commit=False,
        )

        self.assertFalse(review["valid"])
        self.assertFalse(review["committed"])
        self.assertTrue(any("Language must match" in error["error"] for error in review["errors"]))
        self.assertEqual(self.db.get(CollectionItem, self.item.id).lang, "en")


    def test_review_accepts_tag_slab_protection_from_workbook(self):
        data = self._workbook_with_card_edits(N2="tag_slab", Q2="", R2="Pristine 10", S2="TAG-777")

        committed = review_inventory_workbook(
            self.db,
            self.user.id,
            data,
            commit=True,
        )

        self.assertTrue(committed["committed"])
        updated_item = self.db.get(CollectionItem, self.item.id)
        self.assertEqual(updated_item.protection_type, "tag_slab")
        self.assertEqual(updated_item.grader, "TAG")
        self.assertEqual(updated_item.grade, "Pristine 10")
        self.assertEqual(updated_item.certification_number, "TAG-777")

    def test_review_rejects_legacy_condition_labels(self):
        review = review_inventory_workbook(
            self.db,
            self.user.id,
            self._workbook_with_card_edits(H2="Mint"),
            commit=False,
        )

        self.assertFalse(review["valid"])
        self.assertFalse(review["committed"])
        self.assertTrue(any("Condition must be one of: NM, LP, MP, HP, DMG" in error["error"] for error in review["errors"]))
        self.assertEqual(self.db.get(CollectionItem, self.item.id).condition, "NM")


if __name__ == "__main__":
    unittest.main()
