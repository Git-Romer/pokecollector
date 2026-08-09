from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from api.auth import get_current_user
from database import get_db
from services.card_values import effective_market_price, normalize_price_field
from services.card_visibility import visible_card_filter
from models import CollectionItem, Card, ProductPurchase, StorageLocation, User
import io
import csv
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

router = APIRouter()


CARD_HEADERS = [
    "Record UID", "Card ID", "Name", "Set", "Number", "Rarity", "Quantity",
    "Condition", "Variant", "Language", "Cost Basis", "Acquisition Source",
    "Collection Intent", "Protection", "Storage Location UID", "Storage Location",
    "Grading Company", "Grade", "Certification Number", "Grail", "Card History",
    "Notes", "Status", "Added At", "Updated At", "Removal Reason",
]
SEALED_HEADERS = [
    "Record UID", "Product Name", "Product Type", "Acquisition Source", "Collection Intent",
    "Quantity", "Condition", "Cost Basis", "Acquisition Date", "Storage Location UID",
    "Storage Location", "Notes", "Status", "Updated At", "Removal Reason",
]
LOCATION_HEADERS = [
    "Record UID", "Name", "Description", "Default", "Active", "Created At", "Updated At",
]
ERROR_HEADERS = ["Sheet", "Row", "Record UID", "Error"]
EXCEL_CARD_HEADERS = [
    "Card ID", "Name", "Set", "Number", "Rarity", "Quantity", "Condition",
    "Variant", "Language", "Cost Basis", "Collection Intent", "Protection",
    "Grading Company", "Grade", "Certification Number", "Grail",
    "Card History", "Notes", "Status", "Record UID",
]
EXCEL_SEALED_HEADERS = [
    "Product Name", "Product Type", "Quantity", "Condition", "Cost Basis",
    "Acquisition Date", "Collection Intent", "Notes", "Status", "Record UID",
]
EXCEL_ACQUISITION_HEADERS = [
    "Record UID", "Record Type", "Item Name", "Acquisition Source",
    "Acquisition Date", "Storage Location UID", "Storage Location",
    "Storage Type", "Storage Detail", "Updated At",
]


def _date_cell(value):
    if value is None:
        return ""
    return value.isoformat()


def _card_row(item):
    card = item.card
    location = getattr(item, "storage_location", None)
    return [
        item.record_uid,
        card.id,
        card.name,
        card.set_ref.name if card.set_ref else "",
        card.number or "",
        card.rarity or "",
        item.quantity,
        item.condition,
        item.variant,
        item.lang,
        item.purchase_price,
        item.acquisition_source or "",
        getattr(item, "collection_intent", None) or "main_collection",
        item.protection_type or "raw",
        location.record_uid if location else "",
        location.name if location else "",
        item.grader or "",
        item.grade or "",
        item.certification_number or "",
        bool(getattr(item, "is_grail", False)),
        getattr(item, "card_history", None) or "",
        item.notes or "",
        getattr(item, "status", None) or "owned",
        _date_cell(getattr(item, "added_at", None)),
        _date_cell(getattr(item, "updated_at", None)),
        getattr(item, "removal_reason", None) or "",
    ]


def _style_workbook(workbook):
    header_fill = PatternFill("solid", fgColor="111111")
    header_font = Font(color="FFFFFF", bold=True)
    accent_fill = PatternFill("solid", fgColor="00A3E0")

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        sheet.row_dimensions[1].height = 26
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        if sheet.max_column:
            sheet["A1"].fill = accent_fill
        for column in range(1, sheet.max_column + 1):
            values = [
                str(sheet.cell(row=row, column=column).value or "")
                for row in range(1, min(sheet.max_row, 100) + 1)
            ]
            width = min(max(max((len(value) for value in values), default=8) + 2, 11), 34)
            sheet.column_dimensions[get_column_letter(column)].width = width


def build_collection_workbook(items, products, locations=()) -> bytes:
    workbook = Workbook()
    owned = workbook.active
    owned.title = "Owned Cards"
    owned.append(CARD_HEADERS)
    bulk = workbook.create_sheet("Bulk")
    bulk.append(CARD_HEADERS)

    for item in items:
        if not item.card:
            continue
        target = bulk if getattr(item, "inventory_kind", "owned") == "bulk" else owned
        target.append(_card_row(item))

    sealed = workbook.create_sheet("Sealed Products")
    sealed.append(SEALED_HEADERS)
    for product in products:
        location = getattr(product, "storage_location", None)
        sealed.append([
            product.record_uid,
            product.product_name,
            product.product_type or "",
            product.acquisition_source or "",
            getattr(product, "collection_intent", None) or "main_collection",
            product.quantity,
            product.sealed_condition,
            product.purchase_price,
            _date_cell(product.purchase_date),
            location.record_uid if location else "",
            location.name if location else "",
            product.notes or "",
            product.status or "active",
            _date_cell(product.updated_at),
            product.removal_reason or "",
        ])

    location_sheet = workbook.create_sheet("Storage Locations")
    location_sheet.append(LOCATION_HEADERS)
    for location in locations:
        location_sheet.append([
            location.record_uid,
            location.name,
            location.description or "",
            bool(location.is_default),
            bool(location.is_active),
            _date_cell(location.created_at),
            _date_cell(location.updated_at),
        ])

    errors = workbook.create_sheet("Import Errors")
    errors.append(ERROR_HEADERS)
    _style_workbook(workbook)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_excel_export(items, products) -> bytes:
    """Build the collection's three-sheet, non-market Excel export."""
    workbook = Workbook()
    cards = workbook.active
    cards.title = "Cards"
    cards.append(EXCEL_CARD_HEADERS)

    acquisition_rows = []
    for item in items:
        card = item.card
        if not card:
            continue
        location = getattr(item, "storage_location", None)
        cards.append([
            card.id,
            card.name,
            card.set_ref.name if card.set_ref else "",
            card.number or "",
            card.rarity or "",
            item.quantity,
            item.condition,
            item.variant,
            item.lang,
            item.purchase_price,
            getattr(item, "collection_intent", None) or "main_collection",
            item.protection_type or "raw",
            item.grader or "",
            item.grade or "",
            item.certification_number or "",
            bool(getattr(item, "is_grail", False)),
            getattr(item, "card_history", None) or "",
            item.notes or "",
            getattr(item, "status", None) or "owned",
            item.record_uid,
        ])
        acquisition_rows.append([
            item.record_uid,
            "Card",
            card.name,
            item.acquisition_source or "",
            _date_cell(getattr(item, "added_at", None)),
            location.record_uid if location else "",
            location.name if location else "",
            item.storage_type or "",
            item.storage_detail or "",
            _date_cell(getattr(item, "updated_at", None)),
        ])

    sealed = workbook.create_sheet("Sealed Product")
    sealed.append(EXCEL_SEALED_HEADERS)
    for product in products:
        location = getattr(product, "storage_location", None)
        sealed.append([
            product.product_name,
            product.product_type or "",
            product.quantity,
            product.sealed_condition,
            product.purchase_price,
            _date_cell(product.purchase_date),
            getattr(product, "collection_intent", None) or "main_collection",
            product.notes or "",
            product.status or "active",
            product.record_uid,
        ])
        acquisition_rows.append([
            product.record_uid,
            "Sealed Product",
            product.product_name,
            product.acquisition_source or "",
            _date_cell(product.purchase_date),
            location.record_uid if location else "",
            location.name if location else "",
            product.storage_type or "",
            product.storage_detail or "",
            _date_cell(product.updated_at),
        ])

    acquisition = workbook.create_sheet("Acquisition & Storage")
    acquisition.append(EXCEL_ACQUISITION_HEADERS)
    for row in acquisition_rows:
        acquisition.append(row)

    _style_workbook(workbook)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


@router.get("/xlsx")
def export_xlsx(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    items = db.query(CollectionItem).join(Card, Card.id == CollectionItem.card_id).options(
        joinedload(CollectionItem.card).joinedload(Card.set_ref),
        joinedload(CollectionItem.storage_location),
    ).filter(
        CollectionItem.user_id == current_user.id,
        visible_card_filter(db, current_user.id, "all"),
    ).all()
    products = db.query(ProductPurchase).options(
        joinedload(ProductPurchase.storage_location)
    ).filter(
        ProductPurchase.user_id == current_user.id
    ).order_by(ProductPurchase.purchase_date.desc()).all()
    filename = f"john-johns-pc-{datetime.date.today().isoformat()}.xlsx"
    return StreamingResponse(
        io.BytesIO(build_excel_export(items, products)),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _normalize_currency(value: str | None) -> tuple[str, str]:
    currency = (value or "EUR").upper()
    if currency == "USD":
        return "USD", "$"
    return "EUR", "€"


def _convert_eur(amount: float | None, exchange_rate: float, currency: str) -> float | None:
    if amount is None:
        return None
    return float(amount) * exchange_rate if currency == "USD" else float(amount)


def _format_money(amount: float | None, symbol: str) -> str:
    if amount is None:
        return "-"
    return f"{symbol}{amount:.2f}"


@router.get("/csv")
def export_csv(
    price_field: str = Query(default="price_trend", description="Price field to use for value calculation"),
    currency: str = Query(default="EUR", description="Display currency"),
    exchange_rate: float = Query(default=1.0, gt=0, description="EUR to selected currency rate"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export collection as CSV."""
    price_field = normalize_price_field(price_field)
    currency, symbol = _normalize_currency(currency)
    items = db.query(CollectionItem).join(Card, Card.id == CollectionItem.card_id).options(
        joinedload(CollectionItem.card).joinedload(Card.set_ref)
    ).filter(
        CollectionItem.user_id == current_user.id,
        visible_card_filter(db, current_user.id, "all"),
    ).all()

    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        "Card ID", "Name", "Set", "Number", "Rarity",
        "Quantity", "Condition", f"Cost Basis ({currency})",
        f"Current Price ({currency})", f"Total Value ({currency})",
        "Added At"
    ])

    # Rows
    for item in items:
        card = item.card
        if not card:
            continue
        set_name = card.set_ref.name if card.set_ref else ""
        current_price = effective_market_price(card, item.variant, price_field)
        display_current_price = _convert_eur(current_price, exchange_rate, currency)
        display_purchase_price = _convert_eur(item.purchase_price, exchange_rate, currency)
        total_value = round((display_current_price or 0) * item.quantity, 2)

        writer.writerow([
            card.id,
            card.name,
            set_name,
            card.number or "",
            card.rarity or "",
            item.quantity,
            item.condition,
            display_purchase_price or "",
            display_current_price or "",
            total_value,
            item.added_at.strftime("%Y-%m-%d") if item.added_at else "",
        ])

    output.seek(0)
    filename = f"john-johns-pc-{datetime.date.today().isoformat()}.csv"

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/pdf")
def export_pdf(
    price_field: str = Query(default="price_trend", description="Price field to use for value calculation"),
    currency: str = Query(default="EUR", description="Display currency"),
    exchange_rate: float = Query(default=1.0, gt=0, description="EUR to selected currency rate"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export collection as PDF."""
    price_field = normalize_price_field(price_field)
    currency, symbol = _normalize_currency(currency)
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import mm

        items = db.query(CollectionItem).join(Card, Card.id == CollectionItem.card_id).options(
            joinedload(CollectionItem.card).joinedload(Card.set_ref)
        ).filter(
            CollectionItem.user_id == current_user.id,
            visible_card_filter(db, current_user.id, "all"),
        ).all()

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=10*mm,
            leftMargin=10*mm,
            topMargin=10*mm,
            bottomMargin=10*mm,
        )

        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Title"],
            fontSize=18,
            spaceAfter=6,
        )
        story.append(Paragraph("John John's PC", title_style))
        story.append(Paragraph(
            f"Exported: {datetime.date.today().isoformat()} | Total cards: {sum(i.quantity for i in items)}",
            styles["Normal"]
        ))
        story.append(Spacer(1, 10*mm))

        # Table
        headers = ["Name", "Set", "No.", "Rarity", "Qty", "Condition", f"Cost Basis {currency}", f"Current {currency}", f"Value {currency}"]
        data = [headers]

        total_value = 0
        for item in items:
            card = item.card
            if not card:
                continue
            set_name = (card.set_ref.name[:20] if card.set_ref else "")
            current_price = _convert_eur(effective_market_price(card, item.variant, price_field), exchange_rate, currency)
            purchase_price = _convert_eur(item.purchase_price, exchange_rate, currency)
            val = round((current_price or 0) * item.quantity, 2)
            total_value += val

            data.append([
                card.name[:30],
                set_name,
                card.number or "-",
                (card.rarity or "-")[:15],
                str(item.quantity),
                item.condition,
                _format_money(purchase_price, symbol) if purchase_price else "-",
                _format_money(current_price, symbol) if current_price else "-",
                _format_money(val, symbol),
            ])

        # Summary row
        data.append(["", "", "", "", "", "", "", "TOTAL:", _format_money(total_value, symbol)])

        col_widths = [100, 80, 30, 80, 25, 50, 45, 55, 55]
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EE1515")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f5f5f5"), colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#ffe0e0")),
        ]))

        story.append(table)
        doc.build(story)
        buffer.seek(0)

        filename = f"john-johns-pc-{datetime.date.today().isoformat()}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except ImportError:
        return {"error": "reportlab not installed"}
